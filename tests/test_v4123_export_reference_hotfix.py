from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.master_definitions import MASTER_BY_KEY
from core.reporting import safe_excel_sheet_name
from core.selection_labels import reference_record_label

ROOT = Path(__file__).resolve().parents[1]


def test_release_and_build_marker():
    assert (ROOT / "VERSION").read_text().strip() in {"4.12.3", "4.12.4", "4.12.5", "4.12.6", "4.12.7", "4.12.8", "4.12.9", "4.13.0", "4.13.1", "4.13.2", "4.13.3", "4.13.4", "4.13.5", "4.13.6", "4.13.7"}
    assert "4123-SUPPLY-EXPORT-REFERENCE-HOTFIX" in (ROOT / "core/ui.py").read_text()
    assert "4123-SUPPLY-EXPORT-REFERENCE-HOTFIX" in (ROOT / "core/auth.py").read_text()


def test_excel_sheet_name_sanitizes_forbidden_characters_and_length():
    title = "Selected CUSTOMER ORDER EDIT / DELETE: [critical]?*\\test"
    safe = safe_excel_sheet_name(title)
    assert len(safe) <= 31
    for forbidden in "\\/*?:[]":
        assert forbidden not in safe


def test_openpyxl_accepts_sanitized_supply_chain_edit_delete_title():
    frame = pd.DataFrame([{"Order": "450001", "Status": "OPEN"}])
    output = BytesIO()
    safe = safe_excel_sheet_name("Selected CUSTOMER ORDER EDIT / DELETE")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name=safe)
    workbook = load_workbook(BytesIO(output.getvalue()))
    assert len(workbook.sheetnames) == 1
    assert "/" not in workbook.sheetnames[0]
    assert len(workbook.sheetnames[0]) <= 31


def test_supply_chain_records_and_reports_use_shared_excel_sheet_safety():
    for rel in ("app_pages/supply_chain.py", "app_pages/records_center.py", "app_pages/reports.py"):
        assert "safe_excel_sheet_name" in (ROOT / rel).read_text()


def test_reference_master_selector_is_descriptive_for_party_records():
    definition = MASTER_BY_KEY["customers"]
    row = {
        "id": "00000000-0000-0000-0000-000000000001",
        "party_code": "CUST-0001",
        "party_name": "Kessler + Co. GmbH & Co. KG",
        "country": "Germany",
        "city": "Aalen",
        "contact_person": "Quality Manager",
        "approval_status": "APPROVED",
        "status": "ACTIVE",
    }
    label = reference_record_label(definition, row, {})
    assert "CUST-0001" in label
    assert "Kessler + Co. GmbH & Co. KG" in label
    assert "Germany" in label
    assert "Aalen" in label
    assert "APPROVED" in label


def test_reference_master_selector_resolves_lookup_details():
    definition = MASTER_BY_KEY["approved_sources"]
    row = {
        "id": "00000000-0000-0000-0000-000000000010",
        "source_code": "SRC-0007",
        "part_id": "part-1",
        "supplier_id": "supplier-1",
        "steel_mill_id": "mill-1",
        "supplier_part_number": "SUP-PN-44",
        "approval_reference": "PPAP-2026-17",
        "approved": True,
    }
    lookup_maps = {
        "parts": {"part-1": "40256626 · Differential Shaft · Kessler"},
        "suppliers": {"supplier-1": "SUP-0012 · ABC Steel Supplier · India"},
        "steel_mills": {"mill-1": "MILL-0004 · Approved Steel Mill · India"},
    }
    label = reference_record_label(definition, row, lookup_maps)
    assert "SRC-0007" in label
    assert "40256626" in label
    assert "ABC Steel Supplier" in label
    assert "Approved Steel Mill" in label
