from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from core.reporting import metlab_record_pdf_bytes, quality_record_excel_bytes

ROOT = Path(__file__).resolve().parents[1]


def text(rel: str) -> str:
    return (ROOT / rel).read_text(encoding="utf-8")


def test_v4141_version_build_and_additive_heat_contract():
    assert (ROOT / "VERSION").read_text().strip() == "4.14.1"
    marker = "4141-HEAT-SUM-METLAB-TRAVERSE-LOGIN-APPROVAL-PO-PRICE"
    for rel in ("streamlit_app.py", "core/ui.py", "core/auth.py"):
        assert marker in text(rel)
    sql = text("supabase/migrations/20260824144500_qcms_heat_metlab_case_depth_login_approval_price_v4141.sql")
    assert "sum(coalesce(r.certificate_quantity,0))" in sql
    assert "r.id<>new.id" in sql
    assert "combined certified Heat quantity" in sql
    assert "year_format='YY'" in sql
    assert "METLAB_REPORT" in sql


def test_metlab_auto_number_simple_serial_conclusion_and_case_depth_ui():
    page = text("app_pages/metlab_report.py")
    assert 'placeholder="Generated automatically on first save · YY year"' in page
    assert 'disabled=True' in page
    assert '"Sr No": index' in page
    assert "Distance starts at 0.05 mm" in page
    assert 'distances = [0.05]' in page
    assert "Traverse Locations" in page
    assert "Select Existing MetLAB Report to Edit" in page
    assert 'selectbox("Conclusion"' in page
    assert "Accepted Under Reserve" in page
    assert '"conclusion": conclusion' in page


def test_dimensional_direct_edit_and_current_login_approval():
    dim = text("app_pages/dimensional_report.py")
    met = text("app_pages/metlab_report.py")
    service = text("core/inspection_service.py")
    sql = text("supabase/migrations/20260824144500_qcms_heat_metlab_case_depth_login_approval_price_v4141.sql")
    assert "Select Existing Dimensional Report to Edit" in dim
    assert "Enable Controlled Amendment" in dim and "Enable Controlled Amendment" in met
    assert "Approved By (Current Login)" in dim and "Approved By (Current Login)" in met
    assert "employee_for_profile" in service
    assert "qcms_current_login_employee_id" in sql
    assert "Approved By must be the currently logged-in employee" in sql
    assert "qsms_has_module_approve('DIMENSIONAL_REPORT')" in sql
    assert "qsms_has_module_approve('METLAB_REPORT')" in sql


def test_out_of_spec_highlight_metlab_font_scale_and_price_history_print():
    reporting = text("core/reporting.py")
    po = text("core/purchase_order_reporting.py")
    part = text("app_pages/part_master.py")
    supply = text("core/supply_chain_service.py")
    assert "_controlled_styles(scale=1.20)" in reporting
    assert "_inspection_result_grid" in reporting
    assert "#FEE2E2" in reporting and "Helvetica-Bold" in reporting
    assert "CASE DEPTH / MICROHARDNESS TRAVERSE" in reporting
    assert "LinePlot" in reporting
    assert "PRICE REVISION HISTORY" in po
    assert "Remark" in part and '"remarks": str(row.get("Remark")' in part
    assert '"remarks": row.get("remarks")' in supply


def _metlab_payload():
    return {
        "record": {
            "report_number": "MLAB-D9-26-00008",
            "test_date": "2026-08-24",
            "status": "FINAL",
            "overall_result": "HOLD",
            "disposition": "ACCEPTED_UNDER_RESERVE",
            "disposition_reason": "Controlled reserve",
            "remarks": "Local case result reviewed under concession.",
            "heat_number": "H4588",
            "heat_code": "4588",
            "inspection_scope": "FINAL_DISPATCH_STAGE",
        },
        "part": {"part_number": "2731", "part_name": "Bush", "fsi_part_number": "2731"},
        "customer": {"party_name": "K Drive"},
        "supplier": {"party_name": "PM"},
        "material_grade": {"grade_code": "16MnCr5"},
        "results": {
            "conclusion": "ACCEPTED_UNDER_RESERVE",
            "rows": [
                {"parameter": "Surface Hardness", "specification": "55-60 HRC", "actual_value": "58", "unit": "HRC", "result": "PASS"},
                {"parameter": "GBO / Decarb", "specification": "15 Micron max", "actual_value": "22 Micron", "unit": "", "result": "FAIL"},
            ],
            "chemistry_rows": [], "jominy_rows": [], "requirement_rows": [],
            "case_depth_traverse": {
                "locations": ["Ground Face", "ID", "OD"],
                "rows": [
                    {"distance_mm": 0.05, "Ground Face": 760, "ID": 775, "OD": 768},
                    {"distance_mm": 0.10, "Ground Face": 715, "ID": 746, "OD": 728},
                    {"distance_mm": 0.20, "Ground Face": 680, "ID": 690, "OD": 716},
                ],
            },
        },
        "employees": {}, "microstructure_images": [],
    }


def test_metlab_pdf_and_excel_include_controlled_case_depth_and_fail_styling():
    payload = _metlab_payload()
    pdf = metlab_record_pdf_bytes(payload)
    assert pdf.startswith(b"%PDF") and len(pdf) > 5000
    xlsx = quality_record_excel_bytes(payload, "METLAB")
    wb = load_workbook(BytesIO(xlsx))
    assert "Case Depth Traverse" in wb.sheetnames
    ws = wb["MetLAB Results"]
    headers = {cell.value: cell.column for cell in ws[1]}
    fail_row = 3
    assert ws.cell(fail_row, headers["result"]).font.bold is True
    actual_col = headers["actual_value"]
    assert ws.cell(fail_row, actual_col).font.bold is True
    decision = wb["Conclusion and Decision"]
    dh = {cell.value: cell.column for cell in decision[1]}
    assert decision.cell(2, dh["Conclusion"]).value == "ACCEPTED_UNDER_RESERVE"
    assert decision.cell(2, dh["Conclusion"]).font.bold is True
