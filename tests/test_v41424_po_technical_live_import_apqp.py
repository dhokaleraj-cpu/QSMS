from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from core.master_import_templates import build_live_master_import_template

ROOT = Path(__file__).resolve().parents[1]


class FakeRepo:
    def select(self, table, order_by=None, limit=5000, **kwargs):
        data = {
            "parties": [
                {"party_code": "CUS-001", "party_name": "Customer One", "party_types": ["CUSTOMER"], "city": "Pune", "state": "MH", "country": "India", "status": "ACTIVE"},
                {"party_code": "SUP-001", "party_name": "Supplier One", "party_types": ["SUPPLIER"], "city": "Pune", "state": "MH", "country": "India", "status": "ACTIVE"},
                {"party_code": "MIL-001", "party_name": "Steel Mill One", "party_types": ["STEEL_MILL"], "status": "ACTIVE"},
                {"party_code": "OSP-001", "party_name": "OSP One", "party_types": ["OSP_VENDOR"], "status": "ACTIVE"},
            ],
            "parts": [{"part_number": "40257237", "fsi_part_number": "FSI-01", "part_name": "Diff Pin", "status": "ACTIVE"}],
            "material_grades": [{"material_number": "MAT-0008", "grade_code": "42CrMo4", "standard": "EN10083-3", "status": "ACTIVE"}],
            "processes": [{"process_code": "P10", "process_name": "Machining", "process_type": "IN_HOUSE", "status": "ACTIVE"}],
            "inspection_stages": [{"stage_code": "FINAL", "stage_name": "Final Inspection", "sequence_no": 10, "status": "ACTIVE"}],
            "quality_assets": [{"asset_code": "GA-01", "asset_name": "Micrometer", "asset_type": "GAUGE", "status": "ACTIVE"}],
            "employees": [{"employee_code": "EMP-01", "first_name": "Test", "last_name": "User", "email": "test@example.com", "department": "Quality", "designation": "Engineer", "plant": "D9", "status": "ACTIVE"}],
            "customer_standards": [{"standard_code": "STD-01", "standard_name": "Customer Spec", "revision_number": "A", "author_name": "Customer", "status": "ACTIVE"}],
        }
        return data.get(table, [])[:limit]


def test_live_master_import_template_contains_current_reference_data():
    content = build_live_master_import_template(
        FakeRepo(), ROOT / "templates" / "Part_Master_Template.xlsx",
        selected_key="parts", selected_label="Part Master", version="4.14.24",
    )
    wb = load_workbook(BytesIO(content), data_only=False)
    assert "CONTROLLED_MASTER_DATA" in wb.sheetnames
    assert "MASTER_CUSTOMERS" in wb.sheetnames
    assert "MASTER_SUPPLIERS" in wb.sheetnames
    assert "MASTER_PARTS" in wb.sheetnames
    assert "MASTER_MATERIAL_GRADES" in wb.sheetnames
    assert wb["MASTER_CUSTOMERS"]["A2"].value == "CUS-001 · Customer One"
    assert "42CrMo4" in str(wb["MASTER_MATERIAL_GRADES"]["A2"].value)


def test_v41424_source_contracts_present():
    reporting = (ROOT / "core" / "purchase_order_reporting.py").read_text(encoding="utf-8")
    service = (ROOT / "core" / "supply_chain_service.py").read_text(encoding="utf-8")
    master_import = (ROOT / "app_pages" / "master_import.py").read_text(encoding="utf-8")
    apqp = (ROOT / "app_pages" / "npd_apqp.py").read_text(encoding="utf-8")
    assert "RAW MATERIAL DETAILS & SUPPLIER TECHNICAL DATA" in reporting
    assert 'source == "STANDARD"' in reporting and 'source != "CUSTOM"' in reporting
    assert '"source": "CUSTOM"' in service and '"source": "STANDARD"' in service
    assert "Controlled Template + Live Master Data" in master_import
    assert "def _optional_date" in apqp and "overdue = 0" in apqp
