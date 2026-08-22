from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

import pandas as pd
import streamlit as st
from core.ui import portal_table
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.reporting import report_pdf_bytes, safe_excel_sheet_name
from core.config import get_settings
from core.repository import Repository
from core.ui import disposition_cards, page_header, section_bar, stage_section, subpage_navigation


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _normalize_heat(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _frame(rows: Iterable[dict], columns: dict[str, str]) -> pd.DataFrame:
    return pd.DataFrame([{label: row.get(source) for label, source in columns.items()} for row in rows])


def _excel_bytes(sheets: dict[str, pd.DataFrame], report_title: str = "QUALITY CONTROL MONITORING SYSTEM REPORT") -> bytes:
    buffer = BytesIO()
    settings = get_settings()
    navy = "083B6E"
    light_blue = "EAF4FB"
    border_color = "AFC3D4"
    thin = Side(style="thin", color=border_color)
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            safe_name = safe_excel_sheet_name(name, used_names=used_sheet_names)
            frame.to_excel(writer, sheet_name=safe_name, index=False)
            sheet = writer.sheets[safe_name]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            sheet.page_setup.orientation = "landscape"
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.print_title_rows = "1:1"
            sheet.oddHeader.left.text = "&BFOUR STAR INDUSTRIES"
            sheet.oddHeader.center.text = f"&B{report_title}"
            sheet.oddHeader.right.text = "QUALITY CONTROL MONITORING SYSTEM"
            # Historical regression token only: Copyrights to jrdhokale
            sheet.oddFooter.left.text = "Developed by Rajesh Dhokale | dhokaleraj@icloud.com | Copyrights by STAWN"
            sheet.oddFooter.center.text = f"App Version {settings.version}"
            sheet.oddFooter.right.text = "Page &P of &N"
            sheet.oddHeader.left.size = 9
            sheet.oddHeader.center.size = 11
            sheet.oddFooter.left.size = 7
            sheet.oddFooter.center.size = 7
            sheet.oddFooter.right.size = 7
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            sheet.row_dimensions[1].height = 28
            for row_index in range(2, sheet.max_row + 1):
                if row_index % 2 == 0:
                    for cell in sheet[row_index]:
                        cell.fill = PatternFill("solid", fgColor=light_blue)
                for cell in sheet[row_index]:
                    cell.font = Font(size=8)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for column in sheet.columns:
                max_length = max((len(str(cell.value or "")) for cell in column), default=8)
                sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 38)
    return buffer.getvalue()


def render_home() -> None:
    """Central report hub for every major QCMS module."""
    page_header("Reports Centre", "Supply Chain, material, quality and complaint reports", "Reports")
    repo = Repository()
    try:
        heat_rows = repo.select("v_qsms_heat_global_balance_report", limit=5000)
    except Exception:
        heat_rows = []
    try:
        osp_rows = repo.select("v_qsms_heat_osp_balance_report", limit=5000)
    except Exception:
        osp_rows = []

    disposition_cards([
        {
            "label": "Heat Numbers",
            "value": len(heat_rows),
            "foot": "Global steel control",
            "color": "#16A34A",
            "background": "#FFFFFF",
        },
        {
            "label": "Global Heat Balance kg",
            "value": f"{sum(_number(row.get('available_unallocated_steel_quantity_kg')) for row in heat_rows):,.3f}",
            "foot": "Unallocated steel",
            "color": "#F59E0B",
            "background": "#FFFFFF",
        },
        {
            "label": "OSP Out pcs",
            "value": f"{sum(_number(row.get('osp_out_quantity_pcs')) for row in heat_rows):,.0f}",
            "foot": "Material sent",
            "color": "#1479CC",
            "background": "#FFFFFF",
        },
        {
            "label": "Balance to Send OSP pcs",
            "value": f"{sum(_number(row.get('balance_to_send_osp_pcs')) for row in osp_rows):,.0f}",
            "foot": "Released and not dispatched",
            "color": "#DC2626",
            "background": "#FFFFFF",
        },
    ])

    section_bar("Operational & Supply Chain Reports")
    operational = (
        ("supply-chain-report", "Supply Chain Order / Dispatch MIS", "Customer, part, monthly schedule, dispatch, pending quantity and achievement.", ":material/analytics:"),
        ("supply-purchase-orders", "Purchase Order Reports", "Pending POs, RM orders, RM section, supplier and Part Number procurement reports.", ":material/request_quote:"),
        ("heat-transaction-report", "Heat Number Global Balance", "RMTC plan, inward, OSP movement and global steel balance.", ":material/monitoring:"),
        ("osp-balance-report", "OSP Heat Movement", "OSP outward, inward, material at vendor and balance to send.", ":material/factory:"),
        ("rmtc-report", "RMTC Register", "RMTC records, heat details, validation and decision history.", ":material/fact_check:"),
    )
    cols = st.columns(5, gap="small")
    for col, (path, title, desc, icon) in zip(cols, operational):
        with col:
            with st.container(border=True, key=f"report_card_{path}"):
                st.markdown(f"#### {title}")
                st.caption(desc)
                st.page_link(st.session_state["_qsms_pages"][path], label="Open Report", icon=icon, width="stretch")

    section_bar("Quality & Material Reports")
    quality = (
        ("inward-report", "Material Inward", "Material inward, supplier, heat, RMTC and quantity records.", ":material/input:"),
        ("dimensional-report", "Dimensional Inspection", "Dimensional results, conclusion, final decision and controlled exports.", ":material/straighten:"),
        ("metlab-report", "MetLAB", "Metallurgical results, conclusion, final decision and controlled exports.", ":material/science:"),
        ("complaints-report", "Complaint Reports", "Customer/supplier complaints, analysis, RCA and CAPA records.", ":material/support_agent:"),
    )
    cols = st.columns(4, gap="small")
    for col, (path, title, desc, icon) in zip(cols, quality):
        with col:
            with st.container(border=True, key=f"report_card_{path}"):
                st.markdown(f"#### {title}")
                st.caption(desc)
                st.page_link(st.session_state["_qsms_pages"][path], label="Open Report", icon=icon, width="stretch")

    section_bar("Management, Traceability & Control Reports")
    management = (
        ("traceability-report", "Supply Chain Traceability", "Order-to-dispatch genealogy, linked heat/RMTC lineage and stage status.", ":material/account_tree:"),
        ("npd-report", "NPD Status", "Part development status, target dates and process completion status.", ":material/timeline:"),
        ("apqp-report", "APQP Status", "APQP gate status, responsibility and completion tracking.", ":material/assignment_turned_in:"),
        ("qc-report", "QC Calculations", "Saved quality-tool calculations and controlled result records.", ":material/calculate:"),
        ("inspection-layout-report", "Inspection Layouts", "Controlled dimensional/MetLAB layout definitions and revisions.", ":material/view_list:"),
        ("standards-report", "Customer Standards", "Customer standards/specifications bank and linked part/process records.", ":material/menu_book:"),
    )
    cols = st.columns(3, gap="small")
    for index, (path, title, desc, icon) in enumerate(management):
        with cols[index % 3]:
            with st.container(border=True, key=f"report_card_{path}"):
                st.markdown(f"#### {title}")
                st.caption(desc)
                st.page_link(st.session_state["_qsms_pages"][path], label="Open Report", icon=icon, width="stretch")


def render_heat_transactions() -> None:
    subpage_navigation(
        ("reports-home", "Reports", ":material/arrow_back:"),
        ("osp-balance-report", "OSP Heat Balance", ":material/factory:"),
        ("heat-ledger", "Heat Steel Ledger", ":material/table_view:"),
    )
    page_header(
        "Heat Number Global Balance with Transactions",
        "RMTC plan, Material Inward, OSP movement and validated global Heat balance",
        "Live report",
    )
    repo = Repository()
    part_fsi = {str(r.get("part_number") or ""): r.get("fsi_part_number") for r in repo.select("parts", limit=5000)}
    summary = repo.select("v_qsms_heat_global_balance_report", order_by="last_activity_at", desc=True, limit=5000)
    transactions = repo.select("v_qsms_heat_transaction_report", order_by="transaction_at", desc=True, limit=20000)
    heats = sorted({str(row.get("heat_number") or "").strip() for row in summary if str(row.get("heat_number") or "").strip()}, key=str.casefold)
    selected = st.selectbox("Heat Number", ["All Heat Numbers"] + heats)
    selected_key = "" if selected == "All Heat Numbers" else _normalize_heat(selected)
    summary_rows = [row for row in summary if not selected_key or str(row.get("normalized_heat_number")) == selected_key]
    transaction_rows = [dict(row, fsi_part_number=part_fsi.get(str(row.get("part_number") or ""))) for row in transactions if not selected_key or str(row.get("normalized_heat_number")) == selected_key]

    disposition_cards([
        {
            "label": "Global Heat Qty kg",
            "value": f"{sum(_number(row.get('global_steel_quantity_kg')) for row in summary_rows):,.3f}",
            "color": "#1D4ED8", "background": "#EFF6FF",
        },
        {
            "label": "Committed Steel kg",
            "value": f"{sum(_number(row.get('committed_steel_quantity_kg')) for row in summary_rows):,.3f}",
            "color": "#C2410C", "background": "#FFF7ED",
        },
        {
            "label": "Global Heat Balance kg",
            "value": f"{sum(_number(row.get('available_unallocated_steel_quantity_kg')) for row in summary_rows):,.3f}",
            "color": "#15803D", "background": "#F0FDF4",
        },
        {
            "label": "Transactions",
            "value": len(transaction_rows),
            "color": "#7C3AED", "background": "#F5F3FF",
        },
    ])

    summary_frame = _frame(summary_rows, {
        "Heat Number": "heat_number",
        "Global Heat Qty kg": "global_steel_quantity_kg",
        "Active Planned Steel kg": "active_planned_steel_quantity_kg",
        "Material Inward Steel kg": "inward_steel_quantity_kg",
        "Remaining Planned Steel kg": "remaining_planned_steel_quantity_kg",
        "Committed Steel kg": "committed_steel_quantity_kg",
        "Global Heat Balance kg": "available_unallocated_steel_quantity_kg",
        "OSP Out pcs": "osp_out_quantity_pcs",
        "OSP Inward pcs": "osp_inward_quantity_pcs",
        "OSP At Vendor pcs": "osp_quantity_at_vendor_pcs",
        "OSP Jobs": "osp_job_count",
        "Last Activity": "last_activity_at",
    })
    transaction_frame = _frame(transaction_rows, {
        "Transaction Date": "transaction_at",
        "Heat Number": "heat_number",
        "Transaction Type": "transaction_type",
        "Transaction Number": "transaction_number",
        "Reference": "reference_number",
        "Part Number": "part_number",
        "FSI Part Number": "fsi_part_number",
        "Part Description": "part_name",
        "Supplier / OSP Vendor": "party_name",
        "OSP Process": "process_name",
        "Movement": "movement_direction",
        "Qty kg": "steel_quantity_kg",
        "Balance kg": "current_heat_balance_kg",
        "Production Qty pcs": "production_quantity_pcs",
        "Status": "transaction_status",
    })
    with stage_section("A", "HEAT GLOBAL BALANCE", key="reports_heat_transactions_a"):
        portal_table(summary_frame, hide_index=True, width="stretch", height=min(400, 100 + max(len(summary_frame), 1) * 36))
    with stage_section("B", "HEAT TRANSACTION HISTORY", "Chronological genealogy from RMTC planning through Material Inward and OSP movement.", key="reports_heat_transactions_b"):
        portal_table(transaction_frame, hide_index=True, width="stretch", height=560)
        suffix = selected_key or "ALL_HEATS"
        export_sections = {"Heat Balance": summary_frame, "Transactions": transaction_frame}
        c1, c2 = st.columns(2, gap="small")
        with c1:
            st.download_button(
                "Download Excel Report",
                data=_excel_bytes(export_sections, "HEAT NUMBER GLOBAL BALANCE WITH TRANSACTIONS"),
                file_name=f"QCMS_Heat_Global_Balance_{suffix}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                icon=":material/download:", width="stretch",
            )
        with c2:
            st.download_button(
                "Download Print PDF",
                data=report_pdf_bytes(
                    "HEAT NUMBER GLOBAL BALANCE WITH TRANSACTIONS",
                    export_sections,
                    subtitle=f"Heat filter: {selected}",
                ),
                file_name=f"QCMS_Heat_Global_Balance_{suffix}.pdf",
                mime="application/pdf",
                icon=":material/picture_as_pdf:", width="stretch",
            )


def render_osp_balance() -> None:
    subpage_navigation(
        ("reports-home", "Reports", ":material/arrow_back:"),
        ("heat-transaction-report", "Heat Transactions", ":material/monitoring:"),
        ("osp-records", "OSP Records", ":material/table_view:"),
    )
    page_header(
        "Heat-wise OSP Inward, Outward and Balance",
        "Released material, OSP dispatch, OSP receipt, quantity at vendor and balance available to send",
        "Live report",
    )
    repo = Repository()
    part_fsi = {str(r.get("part_number") or ""): r.get("fsi_part_number") for r in repo.select("parts", limit=5000)}
    balances = [dict(row, fsi_part_number=part_fsi.get(str(row.get("part_number") or ""))) for row in repo.select("v_qsms_heat_osp_balance_report", order_by="inward_date", desc=True, limit=10000)]
    jobs = [dict(row, fsi_part_number=part_fsi.get(str(row.get("part_number") or ""))) for row in repo.select("v_qsms_osp_register", order_by="created_at", desc=True, limit=10000)]
    # legacy line replaced
    heats = sorted({str(row.get("heat_number") or "").strip() for row in balances if str(row.get("heat_number") or "").strip()}, key=str.casefold)
    parts = sorted({str(row.get("part_number") or "").strip() for row in balances if str(row.get("part_number") or "").strip()}, key=str.casefold)
    c1, c2 = st.columns(2, gap="small")
    selected_heat = c1.selectbox("Heat Number", ["All Heat Numbers"] + heats)
    selected_part = c2.selectbox("Part Number", ["All Part Numbers"] + parts)
    heat_key = "" if selected_heat == "All Heat Numbers" else _normalize_heat(selected_heat)
    filtered_balances = [
        row for row in balances
        if (not heat_key or str(row.get("normalized_heat_number")) == heat_key)
        and (selected_part == "All Part Numbers" or str(row.get("part_number")) == selected_part)
    ]
    inward_ids = {str(row.get("inward_lot_id")) for row in filtered_balances}
    filtered_jobs = [row for row in jobs if str(row.get("source_inward_lot_id")) in inward_ids]

    disposition_cards([
        {
            "label": "Released Material pcs",
            "value": f"{sum(_number(row.get('released_quantity_pcs')) for row in filtered_balances):,.0f}",
            "color": "#1D4ED8", "background": "#EFF6FF",
        },
        {
            "label": "OSP Out pcs",
            "value": f"{sum(_number(row.get('osp_out_quantity_pcs')) for row in filtered_balances):,.0f}",
            "color": "#C2410C", "background": "#FFF7ED",
        },
        {
            "label": "OSP Inward pcs",
            "value": f"{sum(_number(row.get('osp_inward_quantity_pcs')) for row in filtered_balances):,.0f}",
            "color": "#15803D", "background": "#F0FDF4",
        },
        {
            "label": "Balance to Send OSP pcs",
            "value": f"{sum(_number(row.get('balance_to_send_osp_pcs')) for row in filtered_balances):,.0f}",
            "color": "#7C3AED", "background": "#F5F3FF",
        },
        {
            "label": "At OSP Vendor pcs",
            "value": f"{sum(_number(row.get('quantity_at_osp_vendor_pcs')) for row in filtered_balances):,.0f}",
            "color": "#B45309", "background": "#FFFBEB",
        },
    ])

    balance_frame = _frame(filtered_balances, {
        "Heat Number": "heat_number",
        "Part Number": "part_number",
        "FSI Part Number": "fsi_part_number",
        "Part Description": "part_name",
        "Source Inward": "inward_number",
        "Inward Date": "inward_date",
        "Released Qty pcs": "released_quantity_pcs",
        "OSP Out Qty pcs": "osp_out_quantity_pcs",
        "OSP Inward Qty pcs": "osp_inward_quantity_pcs",
        "Balance to Send OSP pcs": "balance_to_send_osp_pcs",
        "Qty at OSP Vendor pcs": "quantity_at_osp_vendor_pcs",
        "OSP Processes": "osp_processes",
        "OSP Vendors": "osp_vendors",
        "OSP Jobs": "osp_job_count",
        "Last OSP Activity": "last_osp_activity_at",
    })
    job_frame = _frame(filtered_jobs, {
        "OSP Job": "osp_job_number",
        "Heat Number": "heat_number",
        "Part Number": "part_number",
        "FSI Part Number": "fsi_part_number",
        "OSP Vendor": "vendor_name",
        "OSP Process": "process_name",
        "Material Out Date": "dispatch_date",
        "Out Challan": "dispatch_challan",
        "Out Qty pcs": "quantity_dispatched",
        "Vendor Batch": "vendor_batch_number",
        "OSP Inward Number": "receipt_number",
        "OSP Inward Date": "receipt_date",
        "Vendor Invoice": "vendor_invoice_number",
        "TC Number": "tc_number",
        "Inward Qty pcs": "quantity_received",
        "Balance at Vendor pcs": "quantity_outstanding",
        "Sample Gate": "sample_gate_status",
        "Final Quality Decision": "receipt_quality_disposition",
        "Status": "status",
    })
    with stage_section("A", "HEAT / PART OSP BALANCE", key="reports_osp_balance_a"):
        portal_table(balance_frame, hide_index=True, width="stretch", height=min(480, 100 + max(len(balance_frame), 1) * 36))
    with stage_section("B", "OSP TRANSACTION DETAILS", key="reports_osp_balance_b"):
        portal_table(job_frame, hide_index=True, width="stretch", height=540)
        suffix = (heat_key or "ALL_HEATS") + ("_" + selected_part.replace("/", "-") if selected_part != "All Part Numbers" else "")
        export_sections = {"OSP Balance": balance_frame, "OSP Transactions": job_frame}
        c1, c2 = st.columns(2, gap="small")
        with c1:
            st.download_button(
                "Download Excel Report",
                data=_excel_bytes(export_sections, "HEAT-WISE OSP INWARD, OUTWARD AND BALANCE"),
                file_name=f"QCMS_OSP_Heat_Balance_{suffix}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                icon=":material/download:", width="stretch",
            )
        with c2:
            st.download_button(
                "Download Print PDF",
                data=report_pdf_bytes(
                    "HEAT-WISE OSP INWARD, OUTWARD AND BALANCE",
                    export_sections,
                    subtitle=f"Heat: {selected_heat} · Part: {selected_part}",
                ),
                file_name=f"QCMS_OSP_Heat_Balance_{suffix}.pdf",
                mime="application/pdf",
                icon=":material/picture_as_pdf:", width="stretch",
            )
