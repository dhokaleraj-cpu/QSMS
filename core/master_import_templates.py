from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INDEX_FILL = PatternFill("solid", fgColor="F2F4F6")
MASTER_PREFIX = "MASTER_"


def _text(value: Any) -> str:
    return str(value or "").strip()


def _party_has_type(row: Mapping[str, Any], party_type: str) -> bool:
    raw = row.get("party_types") or []
    if isinstance(raw, str):
        values = {token.strip().upper().replace(" ", "_") for token in raw.replace(";", ",").split(",") if token.strip()}
    else:
        values = {str(token or "").strip().upper().replace(" ", "_") for token in raw if str(token or "").strip()}
    return party_type.upper() in values


def _records(repo: Any, table: str, *, order_by: str | None = None, limit: int = 5000) -> list[dict[str, Any]]:
    return [dict(row) for row in repo.select(table, order_by=order_by, limit=limit)]


def _party_rows(parties: Iterable[Mapping[str, Any]], party_type: str) -> list[list[Any]]:
    out: list[list[Any]] = []
    for row in parties:
        if not _party_has_type(row, party_type):
            continue
        code = _text(row.get("party_code")); name = _text(row.get("party_name"))
        exact = " · ".join(v for v in (code, name) if v)
        out.append([exact, code, name, _text(row.get("city")), _text(row.get("state")), _text(row.get("country")), _text(row.get("status"))])
    return out


def controlled_master_reference_sets(repo: Any) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Return exact current QCMS master values used by import templates.

    The first column of each sheet is a display/import value that the importer can
    resolve exactly (CODE · Name where applicable). UUIDs remain internal and are
    never required from the user.
    """
    parties = _records(repo, "parties", order_by="party_name")
    parts = _records(repo, "parts", order_by="part_number")
    grades = _records(repo, "material_grades", order_by="grade_code")
    processes = _records(repo, "processes", order_by="process_code")
    stages = _records(repo, "inspection_stages", order_by="sequence_no")
    assets = _records(repo, "quality_assets", order_by="asset_code")
    employees = _records(repo, "employees", order_by="employee_code")
    standards = _records(repo, "customer_standards", order_by="standard_code")

    part_rows: list[list[Any]] = []
    for row in parts:
        pn = _text(row.get("part_number")); fsi = _text(row.get("fsi_part_number")); name = _text(row.get("part_name"))
        exact = " · ".join(v for v in (pn, f"FSI {fsi}" if fsi else "", name) if v)
        part_rows.append([exact, pn, fsi, name, _text(row.get("drawing_number")), _text(row.get("drawing_revision") or row.get("revision")), _text(row.get("status"))])

    grade_rows: list[list[Any]] = []
    for row in grades:
        material_no = _text(row.get("material_number")); grade = _text(row.get("grade_code"))
        exact = " · ".join(v for v in (material_no, grade) if v)
        grade_rows.append([exact, material_no, grade, _text(row.get("standard")), _text(row.get("status"))])

    process_rows = [[
        " · ".join(v for v in (_text(r.get("process_code")), _text(r.get("process_name"))) if v),
        _text(r.get("process_code")), _text(r.get("process_name")), _text(r.get("process_type")), _text(r.get("status")),
    ] for r in processes]

    stage_rows = [[
        " · ".join(v for v in (_text(r.get("stage_code")), _text(r.get("stage_name"))) if v),
        _text(r.get("stage_code")), _text(r.get("stage_name")), r.get("sequence_no"), _text(r.get("status")),
    ] for r in stages]

    asset_rows = [[
        " · ".join(v for v in (_text(r.get("asset_code")), _text(r.get("asset_name"))) if v),
        _text(r.get("asset_code")), _text(r.get("asset_name")), _text(r.get("asset_type")), _text(r.get("serial_number")), _text(r.get("status")),
    ] for r in assets]

    employee_rows = []
    for r in employees:
        code = _text(r.get("employee_code")); name = " ".join(v for v in (_text(r.get("first_name")), _text(r.get("last_name"))) if v)
        employee_rows.append([" · ".join(v for v in (code, name) if v), code, name, _text(r.get("email")), _text(r.get("department")), _text(r.get("designation")), _text(r.get("plant")), _text(r.get("status"))])

    standard_rows = [[
        " · ".join(v for v in (_text(r.get("standard_code")), _text(r.get("standard_name")), f"Rev {_text(r.get('revision_number'))}" if _text(r.get("revision_number")) else "") if v),
        _text(r.get("standard_code")), _text(r.get("standard_name")), _text(r.get("revision_number")), _text(r.get("author_name")), _text(r.get("status")),
    ] for r in standards]

    return [
        ("MASTER_CUSTOMERS", ["Import Value (Exact)", "Customer Code", "Customer Name", "City", "State", "Country", "Status"], _party_rows(parties, "CUSTOMER")),
        ("MASTER_SUPPLIERS", ["Import Value (Exact)", "Supplier Code", "Supplier Name", "City", "State", "Country", "Status"], _party_rows(parties, "SUPPLIER")),
        ("MASTER_STEEL_MILLS", ["Import Value (Exact)", "Steel Mill Code", "Steel Mill Name", "City", "State", "Country", "Status"], _party_rows(parties, "STEEL_MILL")),
        ("MASTER_OSP_VENDORS", ["Import Value (Exact)", "OSP Vendor Code", "OSP Vendor Name", "City", "State", "Country", "Status"], _party_rows(parties, "OSP_VENDOR")),
        ("MASTER_PARTS", ["Import Value (Exact)", "Part Number", "FSI Part Number", "Part Description", "Drawing Number", "Revision", "Status"], part_rows),
        ("MASTER_MATERIAL_GRADES", ["Import Value (Exact)", "Material Number", "Material Grade", "Standard", "Status"], grade_rows),
        ("MASTER_PROCESSES", ["Import Value (Exact)", "Process Code", "Process Name", "Process Type", "Status"], process_rows),
        ("MASTER_INSPECTION_STAGES", ["Import Value (Exact)", "Stage Code", "Stage Name", "Sequence", "Status"], stage_rows),
        ("MASTER_QUALITY_ASSETS", ["Import Value (Exact)", "Asset Code", "Asset Name", "Asset Type", "Serial Number", "Status"], asset_rows),
        ("MASTER_EMPLOYEES", ["Import Value (Exact)", "Employee Code", "Employee Name", "Email", "Department", "Designation", "Plant", "Status"], employee_rows),
        ("MASTER_CUSTOMER_STANDARDS", ["Import Value (Exact)", "Standard Code", "Standard Name", "Revision", "Issuing Authority", "Status"], standard_rows),
    ]


def _fit_columns(ws: Any, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 10
        for cell in column_cells[:300]:
            value = _text(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def build_live_master_import_template(repo: Any, template_path: str | Path, *, selected_key: str, selected_label: str, version: str) -> bytes:
    """Copy a base import workbook and append live, exact QCMS master-data sheets."""
    path = Path(template_path)
    wb = load_workbook(path)
    for name in list(wb.sheetnames):
        if name.startswith(MASTER_PREFIX) or name == "CONTROLLED_MASTER_DATA":
            del wb[name]

    index = wb.create_sheet("CONTROLLED_MASTER_DATA", 0)
    index["A1"] = "QCMS CONTROLLED IMPORT TEMPLATE — LIVE MASTER DATA"
    index["A1"].font = Font(bold=True, size=14)
    index["A2"] = f"Import Module: {selected_label} ({selected_key})"
    index["A3"] = f"QCMS Version: {version}"
    index["A4"] = f"Generated (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}"
    index["A6"] = "Instructions"
    index["A6"].font = Font(bold=True)
    instructions = [
        "Enter only the variable/new values in the original import sheet(s).",
        "For any field linked to a QCMS master, copy the exact Code / Import Value from the MASTER_* reference sheets.",
        "Do not rename master-reference sheets or change existing controlled master values in this workbook.",
        "If QCMS master data changes, download a fresh template before preparing the next import.",
        "The importer remains duplicate-safe: existing natural keys are skipped and are not silently overwritten.",
    ]
    for idx, line in enumerate(instructions, start=7):
        index.cell(idx, 1, line)
    index["A14"] = "Included Live Reference Sheets"
    index["A14"].font = Font(bold=True)

    sets = controlled_master_reference_sets(repo)
    for row_no, (sheet_name, headers, rows) in enumerate(sets, start=15):
        index.cell(row_no, 1, sheet_name)
        index.cell(row_no, 2, len(rows))
        ws = wb.create_sheet(sheet_name[:31])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64+min(len(headers),26))}1" if len(headers) <= 26 else None
        for col_no, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_no, header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx, value)
        _fit_columns(ws)

    index.column_dimensions["A"].width = 95
    index.column_dimensions["B"].width = 16
    for row in index.iter_rows(min_row=1, max_row=index.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in index[14]:
        cell.fill = INDEX_FILL

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
